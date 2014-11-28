import sys
import os
import pickle
import getopt

from django.core.exceptions import ObjectDoesNotExist

from fasta import fastaManage
from xls import xlsManage

from openpyxl.reader.excel import load_workbook

this_module_path = os.path.abspath (os.path.dirname(sys.modules[__name__].__file__))
server_path = os.path.join(this_module_path, '..', '..')
sys.path.append(server_path)

os.environ['DJANGO_SETTINGS_MODULE'] = 'server.config.settings'

from server.scaffold.models import Scaffold, ScaffoldPosition
from server.scaffold.models import UndefinedScaffold
from server.contig.models import Contig

from server.config import static

# =================================================================================================
# =================================== UTILS =======================================================
# =================================================================================================

## Funkcja pobiera z pliku fasta sekwencje danego scaffolda
#  @param scaff_id: ID poszukiwanego scafolda
#  @param assem: Typ asemblacji. Domyslny typ: CELERA
#  @return: Sekwencja poszukiwanego scaffolda lub None w przypadku nie znalezienia.
def getScaffSeq(scaff_id, assem = fastaManage.CELERA):
    scaffolds = fastaManage.loadScaffolds(assem)    # int(scaff_ID)  : str(sekwencja)
    try:
        return scaffolds[scaff_id]
    except KeyError:
        print "Brak scaffolda", scaff_id, "w pliku fasta!"
        return None

## Funkcja pobiera z pliku fasta sekwencje danego contiga
#  @param cont_id: ID poszukiwanego contiga
#  @param assem: Typ asemblacji. Domyslny typ: CELERA
#  @return: Sekwencja poszukiwanego contiga lub None w przypadku nie znalezienia.
def getContSeq(cont_id, assem = fastaManage.CELERA):
    contigs = fastaManage.loadContigs(assem)    # int(cont_ID)  : str(sekwencja)
    try:
        try:
            return contigs[int(cont_id)]
        except ValueError:
            print "Niepoprawne ID contiga:", cont_id
            return None
    except KeyError:
        print "Brak contiga", cont_id, "w pliku fasta!"
        return None

## Funkcja pobiera z pliku xlsx informacje o danym contigu (przynaleznosc do scaffolda oraz dlugosc sekwencji)
#  @param cont_id: ID poszukiwanego contiga
#  @return: Informacje o znalezionym contigu lub None w przypadku nieznalezienia.
def getContInfoFromXlsx(cont_id):
    # str(cont_ID)  : (str(scaff_ID), str(cont_LEN))
    contigs = xlsManage.getCeleraContigsOnScaffolds()
    try:
        return contigs[str(cont_id)]
    except KeyError:
        print "Brak contiga", cont_id, "w pliku xlsx!"
        return None

# =================================================================================================
# =================================================================================================

## Funkcja odczytuje z pliku xlsx odwzorowanie scaffoldow arachne i celery
#  @param pickle_path: Sciezka do pickla. 
# TODO: Metoda prawdopodobnie niepotrzebna - stad - niedokonczona ;)
def getCeleraArachneMap(pickle_path = None):
    xls_file = load_workbook("xls\\src\\laczenie.xlsx")
    scaffolds = {}
    for sheet in xls_file.worksheets:
        if sheet.title in ['finalne']:
            # Odczytanie pozycji kolumn
            first_row = sheet.rows[0]
            scaff_cel_col = -1
            scaff_ara_col = -1
            for i, cell in enumerate(first_row):
                # Odczytanie ID scaffoldu
                if "scf_celera" in cell.value:
                    scaff_cel_col = i
                    continue
                if "scf_ara_links" in cell.value:
                    scaff_ara_col = i
                    continue
            if scaff_cel_col == -1:
                raise "Nie znaleziono kolumny 'scf_celera'!"
                sys.exit(1)
            if scaff_ara_col == -1:
                raise "Nie znaleziono kolumny 'scf_ara_links'!"
                sys.exit(1)

            # Kolejne scaffoldy
            for row in sheet.rows[1:]:
                # ID celera
                id_cel = str(row[scaff_cel_col].value)
                if not isinstance(id_cel, unicode):
                    id_cel = unicode(id_cel)
                id_cel = id_cel.encode('utf8')

                # Nie mamy jeszcze tego ID celery - odczytujemy odpowiadajacy mu scaffold arachne
                if not scaffolds.has_key(id_cel):    
                    # ID arachne
                    id_ara = str(row[scaff_ara_col].value)
                    if not isinstance(id_ara, unicode):
                        id_ara = unicode(id_ara)
                    id_ara = id_ara.encode('utf8')

                    scaffolds[id_cel] = id_ara

    return scaffolds

# =======================================================================================================================================
# ========================================================= SCAFFOLDY  ==================================================================
# =======================================================================================================================================

## Funkcja dodaje dlugosc sekwencji oraz sama sekwencje do scaffoldow zawartych w bazie danych
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
def addLenAndSeqScaffolds(assemb, debug = False):
    # FASTA plik ( int(ID) : str(SEQ) )
    scaffolds_fasta = fastaManage.loadScaffolds(assemb)
    # XLS plik - dlugosci ( str(ID) : str(LEN) ); zawiera rowniez scaffoldy dzielone (z dopiskiem a lub b)
    scaffolds_xlsx = xlsManage.getScaffoldsLen(assemb, False)

    print "Zapisuje sekwencje oraz dlugosc sekwencji do bazy danych..."
    divided_scaffs = 0
    # Badania dlugosci dla scaffoldow NIE dzielonych
    for sc_id_xlsx, sc_len_xlsx in scaffolds_xlsx.iteritems():
        try:
            # Chcemy na razie tylko te ID calkowitoliczbowe
            sc_id = int(sc_id_xlsx)

            # ID scaffoldu jest w tym momencie liczba calkowita
            if not scaffolds_fasta.has_key(sc_id):
                print "\tID: ", sc_id, "nie znajduje sie w pliku fasta (jego dlugosc:", sc_len_xlsx, ")!"
                continue

            try:
                if int(sc_len_xlsx) != len(scaffolds_fasta[sc_id]):
                    #print "\t\tID:", sc_id, "- rozna wartosc dlugosci sekwencji (", int(sc_len_xlsx), "vs", len(scaffolds_fasta[sc_id]), ")!"
                    #continue
                    pass
            except ValueError:
                print "Niepoprawna dlugosc scaffolda odczytana z pliku xlsx:", sc_len_xlsx
                continue
            
            # Wszystko sie zgadza - uaktualniamy baze
            updateScaffold(sc_id_xlsx, assemb, len(scaffolds_fasta[sc_id]), scaffolds_fasta[sc_id])

        except ValueError:
            if sc_id_xlsx == None or len(str(sc_id_xlsx)) == 0:
                print "Niepoprawne ID:", sc_id_xlsx, "o dlugosci:", sc_len_xlsx
                continue
            try:
                basic_sc_id = int(str(sc_id_xlsx)[:-1])
                divided_scaffs += 1
                print "=== SCAFFOLD DZIELONY  - ", sc_id_xlsx, " - ", basic_sc_id, " - DLUGOSC:", sc_len_xlsx, "==="
                if 'a' in str(sc_id_xlsx):
                    updateScaffold(sc_id_xlsx, assemb, sc_len_xlsx, str(scaffolds_fasta[basic_sc_id])[int(sc_len_xlsx):])
                elif 'b' in str(sc_id_xlsx):
                    updateScaffold(sc_id_xlsx, assemb, sc_len_xlsx, str(scaffolds_fasta[basic_sc_id])[:int(sc_len_xlsx)])

                continue
            except ValueError:
                print "Scaffold", sc_id_xlsx, "jest niepoprawny! Pusta kratka w tabeli!"
                continue

    print "\nScaffoldow dzielonych:", divided_scaffs

## Funkcja uzupelniajaca sekwencje wszystkich scaffoldow z bazy
#  @param assemb: Typ asemblacji
def completeAllSeqsScaffolds(assemb):
    # FASTA plik ( int(ID) : str(SEQ) )
    scaffolds_fasta = fastaManage.loadScaffolds(assemb)
    # Scaffoldy bez sekwencji
    scaffs = Scaffold.objects.filter(assemb_type = assemb, length_bp = 0.0)
    print "Scaffoldow bez sekwencji:", len(scaffs)
    
    for scaff in scaffs:
        try:
            seq = scaffolds_fasta[int(scaff.id)]
            scaff.sequence = str(seq)
            scaff.length_bp = float(len(seq))
            scaff.save()
        except KeyError:
            print "Niewlasciwe ID scaffoldu:", scaff.id

## Funkcja uaktualnia dane o scaffoldzie w bazie danych
#  @param sc_id: ID modyfikowanego scaffoldu
#  @param sc_len: Dlugosc sekwencji modyfikowanego scaffoldu
#  @param sc_seq: Sekwencja modyfikowanego scaffoldu
#  @return: 0 w przypadku powodzenia, 1 w przypadku napotkania bledow.
def updateScaffold(sc_id, assemb, sc_len, sc_seq):
    try:
        scaff = Scaffold.objects.get(id = str(sc_id), assemb_type = assemb)
    except ObjectDoesNotExist:
        print "Objekt o ID:", sc_id, "nie istnieje w bazie!"
        return 1

    scaff.sequence = str(sc_seq)
    scaff.length_bp = float(sc_len)
    scaff.save()
    return 0

# =======================================================================================================================================
# =================================================== SCAFFOLDY CELERA ==================================================================
# =======================================================================================================================================

## Funkcja tworzaca scaffoldy celera na podstawie pliku laczenie i arkuszy chr1_a...chr1_n.
#  @param pickle_path: Sciezka do pickla. Jezeli jest podana tworzy plik pickle: xls\\pickle\\scaffolds_celera.pickle
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Zwraca umieszczenie scaffoldow celera na chromosomach (slownik) wraz z pozycjami kolejnych scaffoldow
def createCeleraScaffolds(pickle_path = None, debug = False):
    chromosomes = xlsManage.getCeleraScaffolds(debug)

    # Zrzucenie do pliku
    if pickle_path == None:
        pickle_path = "xls\\pickle\\scaffolds_celera.pickle"
    pickle.dump(chromosomes, open(pickle_path, "wb"))
    return chromosomes

## Funkcja wypisuje na standardowe wyjscie kolejne scaffoldy z pliku laczenie
def printCeleraScaffoldsFromXLSX():
    xls_file = load_workbook("xls\\src\\laczenie.xlsx")

    for sheet in xls_file.worksheets:
        if sheet.title in ['chr1_a', 'chr2_a', 'chr3_a', 'chr4_a', 'chr5_a', 'chr6_a', 'chr7_a']:
            print "Arkusz: ", sheet.title

            # Odczytanie pozycji kolumn
            first_row = sheet.rows[0]
            scaff_id_col = -1
            scaff_pos_col = -1
            for i, cell in enumerate(first_row):
                # Odczytanie ID scaffoldu
                if "scf_celera" in cell.value:
                    scaff_id_col = i
                    continue
                if "cM" in cell.value:
                    scaff_pos_col = i
                    continue
            if scaff_id_col == -1:
                print "Nie znaleziono kolumny 'scf_celera'!"
                sys.exit(1)
            if scaff_pos_col == -1:
                print "Nie znaleziono kolumny 'cM'!"
                sys.exit(1)

            # Kolejne scaffoldy
            for row in sheet.rows[1:]:
                print row[scaff_id_col].value, row[scaff_pos_col].value

## Funkcja zwraca slownik z chromosomami i pozycjami scaffoldow (ewentualnie tworzy plik pickle)
#  @param pickle_path: Sciezka do pickla. 
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Slownik z chromosomami i pozycjami scaffoldow
def getOrCreateCeleraScaffolds(pickle_path = None, debug = False):
    chromosomes = {}
    if pickle_path == None:
        print "Tworze plik pickle..."
        chromosomes = createCeleraScaffolds(pickle_path, debug)
    else:
        if not os.path.exists(pickle_path):
            print "Podany plik pickle nie istnieje!"
            print "Tworze plik pickle..."
            chromosomes = createCeleraScaffolds(pickle_path)
        else:
            print "Odczytuje plik pickle..."
            chromosomes = pickle.load(open(pickle_path, "rb"))

    if debug == True:
        for chr_id, scaffolds_dict in chromosomes.iteritems():
            print "\n>> Chromosom", chr_id, "\n"
            for scaff_id, scaff_poss_frag in scaffolds_dict.iteritems():
                print "\t", scaff_id
                for scaff_poss in scaff_poss_frag:
                    print "\t\t", scaff_poss

    return chromosomes

## Funkcja zapisuje scaffoldy celery w bazie danych (na podstawie pliku pickle - jezeli nie istnieje - tworzy go)
#  @param pickle_path: Sciezka do pickla. 
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
def saveCeleraScaffoldsInDatabase(pickle_path = None, debug = False):
    chromosomes = getOrCreateCeleraScaffolds(pickle_path, debug)

    # Zapis do bazy
    for chr_id, scaffolds_dict in chromosomes.iteritems():
        print "\n>> Zapisuje chromosom", chr_id
        for scaff_id, scaff_poss_frag in scaffolds_dict.iteritems():
            scaff = Scaffold(id = scaff_id, chromosome_id = chr_id, sequence = "", assemb_type = 1, length_bp = 0.0)
            scaff.save()
            for scaff_poss in scaff_poss_frag:
                scaff_pos_obj = ScaffoldPosition(scaff_id = scaff.id, start = scaff_poss[1][0], end = scaff_poss[1][-1], order = scaff_poss[0])
                scaff_pos_obj.save()

    # Zapis sekwencji i dlugosci sekwencji
    addLenAndSeqScaffolds(fastaManage.CELERA, debug)

# =======================================================================================================================================
# ================================================ SCAFFOLDY ARACHNE ====================================================================
# =======================================================================================================================================

## Funkcja tworzaca scaffoldy arachne na podstawie pliku Markery2.xlsx.
#  @param pickle_path: Sciezka do pickla. Jezeli jest podana tworzy plik pickle: xls\\pickle\\scaffolds_arachne.pickle
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Zwraca umieszczenie scaffoldow arachne na chromosomach (slownik) wraz z pozycjami kolejnych scaffoldow
def createArachneScaffolds(pickle_path = None, debug = False):
    chromosomes = xlsManage.getArachneScaffolds(debug)

    # Zrzucenie do pliku
    if pickle_path == None:
        pickle_path = "xls\\pickle\\scaffolds_arachne.pickle"
    pickle.dump(chromosomes, open(pickle_path, "wb"))
    return chromosomes

## Funkcja zwraca slownik z chromosomami i pozycjami scaffoldow (ewentualnie tworzy plik pickle)
#  @param pickle_path: Sciezka do pickla. 
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Slownik z chromosomami i pozycjami scaffoldow
def getOrCreateArachneScaffolds(pickle_path = None, debug = False):
    chromosomes = {}
    if pickle_path == None:
        print "Tworze plik pickle..."
        chromosomes = createArachneScaffolds(pickle_path, debug)
    else:
        if not os.path.exists(pickle_path):
            print "Podany plik pickle nie istnieje!"
            print "Tworze plik pickle..."
            chromosomes = createArachneScaffolds(pickle_path)
        else:
            print "Odczytuje plik pickle..."
            chromosomes = pickle.load(open(pickle_path, "rb"))

    if debug == True:
        for chr_id, scaffolds_dict in chromosomes.iteritems():
            print "\n>> Chromosom", chr_id, "\n"
            for scaff_id, scaff_poss_frag in scaffolds_dict.iteritems():
                print "\t", scaff_id
                for scaff_poss in scaff_poss_frag:
                    print "\t\t", scaff_poss

    return chromosomes

## Funkcja zapisuje scaffoldy arachne w bazie danych (na podstawie pliku pickle - jezeli nie istnieje - tworzy go)
#  @param pickle_path: Sciezka do pickla. 
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
def saveArachneScaffoldsInDatabase(pickle_path = None, debug = False):
    chromosomes = getOrCreateArachneScaffolds(pickle_path, debug)

    # Zapis do bazy
    for chr_id, scaffolds_dict in chromosomes.iteritems():
        print "\n>> Zapisuje chromosom", chr_id, ". Scaffoldow:", len(scaffolds_dict)
        for scaff_id, scaff_poss_frag in scaffolds_dict.iteritems():
            scaff = Scaffold(id = scaff_id, chromosome_id = chr_id, sequence = "", assemb_type = 0, length_bp = 0.0)
            scaff.save()
            for scaff_poss in scaff_poss_frag:
                scaff_pos_obj = ScaffoldPosition(scaff_id = scaff.id, start = scaff_poss[1][0], end = scaff_poss[1][-1], order = scaff_poss[0])
                scaff_pos_obj.save()

    # Zapis sekwencji i dlugosci sekwencji
    #addLenAndSeqScaffolds(fastaManage.ARACHNE, debug)
    completeAllSeqsScaffolds(fastaManage.ARACHNE)

## Funkcja sprawdza czy wszystkie scaffoldy ARACHNE z pliku xlsx maja sekwencje w pliku FASTA
def checkScaffoldsSeqs():
    # FASTA plik ( int(ID) : str(SEQ) )
    scaffolds_fasta = fastaManage.loadScaffolds(fastaManage.ARACHNE)
    # XLS plik - dlugosci ( str(ID) : str(LEN) );
    scaffolds_xlsx = xlsManage.getScaffoldsLen(fastaManage.ARACHNE, False)
    
    for sc_id_xlsx, sc_len_xlsx in scaffolds_xlsx.iteritems():
        try:
            sc_id_xlsx = int(sc_id_xlsx)    # Zakladamy ze sa same calkowitoliczbowe
        except ValueError:
            print "Niepoprawne ID:", sc_id_xlsx
            continue
        
        # sc_id_xlsx musi byc calkowitoliczbowe tutaj
        if not scaffolds_fasta.has_key(sc_id_xlsx):
            print "Brak sekwencji dla scaffoldu:", sc_id_xlsx, sc_len_xlsx
        

# =======================================================================================================================================
# ========================================== NIEZDEFINIOWANE SCAFFOLDY CELERA ===========================================================
# =======================================================================================================================================

## Funkcja zwraca tablice ID scaffoldow na chromosomach
#  @param chromosomes: Slownik opisujacy chromosomy.
#  @return: Tablica ID scaffoldow na chromosomach.
def createScaffoldsIDsFromChromosomes(chromosomes):
    scaffolds_IDs = []
    for chr_id, scaffolds_dict in chromosomes.iteritems():
        #print "W chr", chr_id, "jest ich", len(scaffolds_dict)
        for scaff_id, _ in scaffolds_dict.iteritems():
            if not str(scaff_id).isdigit():
                scaff_id = str(scaff_id)[:-1]
            if not scaff_id in scaffolds_IDs:
                scaffolds_IDs.append(int(scaff_id))

    return scaffolds_IDs

## Funkcja zapisuje niezdefiniowane scaffoldy celera w bazie danych
#  @param pickle_scaff_path: Sciezka do pickla ze scaffoldami. 
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
def saveCeleraUndefinedScaffoldsInDatabase(pickle_scaff_path = None, debug = False):
    # XLSX
    chromosomes_xlsx = getOrCreateCeleraScaffolds(pickle_scaff_path, debug)
    scaffolds_IDs_xlsx = createScaffoldsIDsFromChromosomes(chromosomes_xlsx)

    # FASTA
    scaffolds_fasta = fastaManage.loadScaffolds(fastaManage.CELERA)

    # Wlasciwe odkrywanie niezdefiniowanych scaffoldow
    udef_count = 0
    print ""
    for scaff_fasta_id, scaff_fasta_seq in scaffolds_fasta.iteritems():
        if not scaff_fasta_id in scaffolds_IDs_xlsx:
            #print scaff_fasta_id
            udef_count += 1
            if debug:
                print "\rNiezdefiniowanych scaffoldow:", udef_count,
            undef_scaff = UndefinedScaffold(udef_count, scaff_fasta_id, str(scaff_fasta_seq), len(scaff_fasta_seq), static.CELERA)
            undef_scaff.save()
    print ""

    print "\n", scaffolds_IDs_xlsx
    print "Z XLSX unikatowych:", len(scaffolds_IDs_xlsx)

# =======================================================================================================================================
# ========================================== NIEZDEFINIOWANE SCAFFOLDY ARACHNE ==========================================================
# =======================================================================================================================================



# =======================================================================================================================================
# ================================================== CONTIGI CELERA =====================================================================
# =======================================================================================================================================

## Pomocnicza klasa Contiga do przechowywania tymczasowych wartosci - bez zapisywania ich w bazie
class ContigTemp:
    def __init__(self, cont_id = -1, scaff_id = -1, start = -1, stop = -1, sequence = "", length = -1):
        self.cont_id = cont_id
        self.scaff_id = scaff_id
        self.start = start
        self.stop = stop
        self.sequence = sequence
        self.length = length

# --------- FUNKCJE POBIERAJACE INFORMACJE O CONTIGACH POPRZEZ PRZESZUKANIE WYLACZNIE PLIKOW FASTA -----------

## Funkcja poszukuje w plikach fasta scaffoldow sekwencje contigow
#  @return: Zwraca dwa obiekty: liste znalezionych contigow (ContigTemp) oraz liste ID nieznalezionych contigow.
def getContigsCeleraFasta():
    if not os.path.exists("pickle\\contigs_with_pos.pickle") or not os.path.exists("pickle\\contigs_with_no_pos.pickle"):
        scaffolds = fastaManage.loadScaffolds(fastaManage.CELERA)
        contigs = fastaManage.loadContigs(fastaManage.CELERA)

        print "\nRozpoczynam wyszukiwanie", len(contigs), "contigow w", len(scaffolds), "scaffoldach...\n"

        # Szukamy kazdego z contigow na scaffoldach
        contigs_found = []
        contigs_forgot = []
        i = 0
        for cont_id, cont_seq in contigs.iteritems():
            print "\rContig:", i, "\tZnalezionych:", len(contigs_found), "\tNIEznalezionych:", len(contigs_forgot),
            i += 1
            # Kazdy z contigow przechodzi przez wszystkie scaffoldy i szuka siebie w nich
            for scaff_id, scaff_seq in scaffolds.iteritems():
                if cont_seq in scaff_seq:
                    # Mamy naszego contiga - sprawdzamy jego dokladne umiejscowienie
                    start = scaff_seq.index(cont_seq)
                    stop = start + len(cont_seq)
                    contigs_found.append(ContigTemp(cont_id, scaff_id, start, stop, cont_seq, len(cont_seq)))
                    break
            # Jezeli nie znajdzie sam siebie to dodaje sie do listy nieznalezionych
            else:
                contigs_forgot.append(cont_id)

        print "\n\nZapisuje znalezione informacje w plikach pickle..."
        pickle.dump(contigs_found, open("pickle\\contigs_with_pos.pickle", "wb"))
        pickle.dump(contigs_forgot, open("pickle\\contigs_with_no_pos.pickle", "wb"))
    else:
        print "\nOdczytuje szczegolowe informacje o contigach z plikow pickle..."
        contigs_found = pickle.load(open("pickle\\contigs_with_pos.pickle", "rb"))
        contigs_forgot = pickle.load(open("pickle\\contigs_with_no_pos.pickle", "rb"))

    print"\nDane zapisane/odczytane prawidlowo z pliku fasta. Podsumowanie:"
    print "Znalezionych + NIEznalezionych contigow:", (len(contigs_found) + len(contigs_forgot))
    print "Znalezionych     contigow:", len(contigs_found)
    print "NIEznalezioncych contigow:", len(contigs_forgot)

    return contigs_found, contigs_forgot

## Funkcja sprawdza poprawnosc przyporzadkowania contigow do scaffoldow.
#  Pobiera ona przyporzadkowania z pliku fasta (reczne dopasowanie), zas weryfikuje z informacjami z pliku xlsx.
#  @return: Lista blednie przyporzadkowanych contigow.
def checkCeleraContigs():
    contigs_found, contigs_forgot = getContigsCeleraFasta()
    conts_scaffs_map = xlsManage.getCeleraContigsOnScaffolds()

    not_found = 0
    found = 0
    good_len = 0
    good_scaff = 0
    bad_assig = []
    for cont in contigs_found:
        if conts_scaffs_map.has_key(str(cont.cont_id)):
            found += 1
            if str(cont.length) == conts_scaffs_map[str(cont.cont_id)][1]:
                good_len += 1
            if str(cont.scaff_id) == conts_scaffs_map[str(cont.cont_id)][0]:
                good_scaff += 1
            else:
                bad_assig.append(cont)
        else:
            not_found += 1
    print "\rZnaleziono", found, "(nie znaleziono: ", not_found, ") contigow z prawidlowymi dlugosciami:", good_len, "oraz przyporzadkowaniem:", good_scaff, ".",

    checkCeleraContigsAssign(bad_assig)

    return bad_assig

## Funkcja sprawdza, czy blednie przyporzadkowane contigi (podczas sprawadzania wylacznie plikow fasta) 
#  sa poprawnie przyporzadkowane w pliku xlsx.
#  Ponizsza funkcja dla kazdego blednie przyporzadkowanego contiga sprawdza do jakiego scaffolda zostal 
#  on przypisany (w xlsx). Nastepnie weryfikuje czy sekwencja danego contiga zawiera sie w sekwencji 
#  przypisanego scaffolda (w fasta).
#  @param bad_assig: Lista blednie przyporzadkowanych contigow
def checkCeleraContigsAssign(bad_assig):
    conts_xls = xlsManage.getCeleraContigsOnScaffolds()             # str(cont_ID)  : (str(scaff_ID), str(cont_LEN))
    conts_fasta = fastaManage.loadContigs(fastaManage.CELERA)       # int(cont_ID)  : str(sekwencja)
    scaffs_fasta = fastaManage.loadScaffolds(fastaManage.CELERA)    # int(scaff_ID) : str(sekwencja)

    good_scaff = 0
    bad_scaff = 0

    print "\nSprawdzam czy przyporzadkowanie contigow zgodnie z plikiem laczenie.xlsx jest poprawne - dla contigow o blednych przyporzadkowaniach na podstawie plikow fasta..."
    print "Contigow z blednym przyporzadkowaniem na podstawie plikow fasta:", len(bad_assig)
    # Sprawdzamy kazdego blednie przyporzadkowanego (przy pomocy wylacznie plikow fasta) contiga
    for bad in bad_assig:
        cont_bad_id = bad.cont_id                        # ID aktualnie sprawdzanego contiga
        scaff_id_xls = conts_xls[str(cont_bad_id)][0]    # przyporzadkowane ID scaffoldu pobrane z pliku laczenie.xlsx

        # Sprawdzenie czy w sekwencji scaffoldu o ID = scaff_id_xlsx znajduje sie sekwencja contiga
        cont_seq = str(conts_fasta[int(cont_bad_id)])       # sekwencja contiga
        try:
            scaff_seq = str(scaffs_fasta[int(scaff_id_xls)])    # sekwencja scaffoldu
        except ValueError:
            scaff_seq = str(scaffs_fasta[int(str(scaff_id_xls)[:-1])])    # sekwencja scaffoldu dzielonego
        if cont_seq in scaff_seq:
            good_scaff += 1
        else:
            bad_scaff += 1

        print "\rPoprawnych:", good_scaff, "\tNiepoprawnych:", bad_scaff,


# --------- FUNKCJE POBIERAJACE INFORMACJE O CONTIGACH POPRZEZ PRZESZUKANIE PLIKOW FASTA WYKORZYSTUJAC PLIK LACZENIE.XLSX -----------
# Informacje o przyporzadkowaniu contiga do scaffolda pobierane sa z pliku laczenie.xlsx - sa to sprawdzone informacje

## Funckja dla kazdego contiga odczytanego z pliku xlsx przeszukuje wszystkie sekwencje scaffoldow fasta z proba odnalezienia 
#  sekwencji fasta danego contiga.
#  @return: Zwraca dwa obiekty: liste znalezionych contigow (ContigTemp) oraz liste ID nieznalezionych contigow.
def getContigsCeleraFastaAndXLSX():
    # Szukamy kazdego z contigow z xlsx na scaffoldach
    contigs = []
    bad_contigs = []

    if not os.path.exists("pickle\\contigs_celera_fasta_xls.pickle") or not os.path.exists("pickle\\contigs_bad_celera_fasta_xls.pickle"):
        conts_xls = xlsManage.getCeleraContigsOnScaffolds()             # str(cont_ID)  : (str(scaff_ID), str(cont_LEN))
        conts_fasta = fastaManage.loadContigs(fastaManage.CELERA)       # int(cont_ID)  : str(sekwencja)
        scaffs_fasta = fastaManage.loadScaffolds(fastaManage.CELERA)    # int(scaff_ID) : str(sekwencja)
        i = 0

        print "\nRozpoczynam wyszukiwanie contigow na scaffoldach wg danych z pliku laczenie.xlsx"
        print "Contigow w XLSX do znalezienia w scaffoldach FASTA:", len(conts_xls)
        print "Contigow w FASTA:", len(conts_fasta)
        print "Scaffoldow w FASTA:", len(scaffs_fasta)
        for cont_xls_id, cont_xls_tup in conts_xls.iteritems():
            i += 1
            print "\rContig:", i,

            # ID contiga z pliku xlsx
            try:
                cont_id = int(cont_xls_id)
            except ValueError:
                print "Niepoprawny wpis w pliku laczenie.xlsx. Bledna komorka - pusta!"
                continue

            # ID scaffoldu z pliku xlsx
            try:
                scaff_id = int(cont_xls_tup[0])
            except ValueError:
                scaff_id = int(str(cont_xls_tup[0])[:-1])

            # Czy contig o danym ID jest wyrozniony w pliku fasta i ma sekwencje?
            if not conts_fasta.has_key(cont_id):
                print "Contig", cont_id, "nie znaleziony w cont_celera.fasta! Brak sekwencji!"
                continue
            else:
                cont_seq = conts_fasta[cont_id]

            # Czy scaffold o danym ID na ktorym lezy opisywany contig (wg xlsx) jest w pliku fasta i ma sekwencje?
            if not scaffs_fasta.has_key(scaff_id):
                print "Scaffold", scaff_id, "nie znaleziony w scaff_celera.fasta!"
                continue
            else:
                scaff_seq = scaffs_fasta[scaff_id]

            # Mamy informacje z xlsx, sekwencje contiga oraz sekwencje scaffoldu na ktorym lezy contig
            # Upewniamy sie ze na pewno contig jest wlasnie na tym scaffoldzie
            if not cont_seq in scaff_seq:
                print "Contig", cont_id, "w rzeczywistosci nie jest na scaffoldzie", scaff_id
                bad_contigs.append(int(cont_xls_id))    # ups...nie jest
                continue

            # Mamy go na pewno - tworzymy tymczasowe obiekty
            start = scaff_seq.index(cont_seq)
            stop = start + len(cont_seq)
            contigs.append(ContigTemp(cont_id, scaff_id, start, stop, cont_seq, len(cont_seq)))

        print "\n\nZapisuje znalezione informacje w pliku pickle..."
        pickle.dump(contigs, open("pickle\\contigs_celera_fasta_xls.pickle", "wb"))
        pickle.dump(bad_contigs, open("pickle\\contigs_bad_celera_fasta_xls.pickle", "wb"))
    else:
        print "\nOdczytuje szczegolowe informacje o contigach z pliku pickle..."
        contigs = pickle.load(open("pickle\\contigs_celera_fasta_xls.pickle", "rb"))
        bad_contigs = pickle.load(open("pickle\\contigs_bad_celera_fasta_xls.pickle", "rb"))

    print "Tymczasowa baza zawiera", len(contigs), "contigow. Nie znaleziono:", len(bad_contigs), "contigow."

    return contigs, bad_contigs

## Funkcja przeszukuje sekwencje fasta scaffoldow w poszukiwaniu sekwencji contiga o ID = cont_id
#  @param cont_id: ID poszukiwanego contiga
#  @return: tablica ID scaffoldow w przypadku znalezienia, None w przypadku nieznalezienia
def findCeleraContigInFastaScaffolds(cont_id):
    scaffolds = fastaManage.loadScaffolds(fastaManage.CELERA)
    contigs = fastaManage.loadContigs(fastaManage.CELERA)
    cont_seq = contigs[int(cont_id)]
    found = False
    
    scaffs = []
    
    print "\nRozpoczynam wyszukiwanie contigu", cont_id, "w", len(scaffolds), "scaffoldach...\n"

    for scaff_id, scaff_seq in scaffolds.iteritems():
        if cont_seq in scaff_seq:
            print "Contig:", cont_id, "znaleziono na scaffoldzie:", scaff_id
            found = True
            scaffs.append(scaff_id)
            
    if found == False:
        print "Nie znaleziono contiga:", cont_id
        return None
    
    return scaffs

##
#
#
def saveCeleraContigsInDatabase(debug = False):
    contigs, _ = getContigsCeleraFastaAndXLSX()
    
    contigs.sort(key = lambda x: x.start, reverse = False)
    contigs.sort(key = lambda x: x.cont_id, reverse = False)
    
    print "Ilosc contigow do zapisania:", len(contigs)
    
    for cont in contigs:
        ctg = Contig(id = cont.cont_id, 
                     scaff_id = cont.scaff_id,
                     order = -1,
                     start = cont.start,
                     end = cont.stop,
                     sequence = cont.sequence,
                     length_bp = cont.length)
        ctg.save()

# =======================================================================================================================================
# ================================================== CONTIGI ARACHNE ====================================================================
# =======================================================================================================================================

def getContigsArachneFastaAndXLSX():
    # Szukamy kazdego z contigow z xlsx na scaffoldach
    contigs = []
    bad_contigs = []

    if not os.path.exists("pickle\\contigs_arachne_fasta_xls.pickle") or not os.path.exists("pickle\\contigs_bad_arachne_fasta_xls.pickle"):
        conts_xls = xlsManage.getArachneContigsOnScaffolds()                # int(cont_ID)  : (str(scaff_ID), int(start), int(end), int(cont_LEN))
        conts_fasta = fastaManage.loadContigs(fastaManage.ARACHNE)          # int(cont_ID)  : str(sekwencja)
        scaffs_fasta = fastaManage.loadScaffolds(fastaManage.ARACHNE)       # int(scaff_ID) : str(sekwencja)
        i = 0

        print "\nRozpoczynam wyszukiwanie contigow na scaffoldach wg danych z pliku new_data_arachne.xlsx"
        print "Contigow w XLSX do znalezienia w scaffoldach FASTA:", len(conts_xls)
        print "Contigow w FASTA:", len(conts_fasta)
        print "Scaffoldow w FASTA:", len(scaffs_fasta)
        for cont_xls_id, cont_xls_tup in conts_xls.iteritems():
            i += 1
            print "\rContig:", i,

            # ID contiga z pliku xlsx
            try:
                cont_id = int(cont_xls_id)
            except ValueError:
                print "Niepoprawny wpis w pliku new_data_arachne.xlsx. Bledna komorka - pusta!"
                continue

            # ID scaffoldu z pliku xlsx
            try:
                scaff_id = int(cont_xls_tup[0])
            except ValueError:
                scaff_id = int(str(cont_xls_tup[0])[:-1])

            # Czy contig o danym ID jest wyrozniony w pliku fasta i ma sekwencje?
            if not conts_fasta.has_key(cont_id):
                print "Contig", cont_id, "nie znaleziony w cont_arachne.fasta! Brak sekwencji!"
                continue
            else:
                cont_seq = conts_fasta[cont_id]

            # Czy scaffold o danym ID na ktorym lezy opisywany contig (wg xlsx) jest w pliku fasta i ma sekwencje?
            if not scaffs_fasta.has_key(scaff_id):
                print "Scaffold", scaff_id, "nie znaleziony w scaff_arachne.fasta!"
                continue
            else:
                scaff_seq = scaffs_fasta[scaff_id]

            # Mamy informacje z xlsx, sekwencje contiga oraz sekwencje scaffoldu na ktorym lezy contig
            # Upewniamy sie ze na pewno contig jest wlasnie na tym scaffoldzie
            if not cont_seq in scaff_seq:
                print "Contig", cont_id, "w rzeczywistosci nie jest na scaffoldzie", scaff_id
                bad_contigs.append(int(cont_xls_id))    # ups...nie jest
                continue

            # Mamy go na pewno - tworzymy tymczasowe obiekty
            start = scaff_seq.index(cont_seq)
            stop = start + len(cont_seq)
            contigs.append(ContigTemp(cont_id, scaff_id, start, stop, cont_seq, len(cont_seq)))
        
        print "\n\nZapisuje znalezione informacje w pliku pickle..."
        pickle.dump(contigs, open("pickle\\contigs_arachne_fasta_xls.pickle", "wb"))
        pickle.dump(bad_contigs, open("pickle\\contigs_bad_arachne_fasta_xls.pickle", "wb"))
    else:
        print "\nOdczytuje szczegolowe informacje o contigach z pliku pickle..."
        contigs = pickle.load(open("pickle\\contigs_arachne_fasta_xls.pickle", "rb"))
        bad_contigs = pickle.load(open("pickle\\contigs_bad_arachne_fasta_xls.pickle", "rb"))

    print "Tymczasowa baza zawiera", len(contigs), "contigow. Nie znaleziono:", len(bad_contigs), "contigow."

    return contigs, bad_contigs

def getContigsArachneFastaAndXLSX_2():
    # Szukamy kazdego z contigow z xlsx na scaffoldach
    contigs = []
    bad_contigs = []

    if not os.path.exists("pickle\\contigs_arachne_fasta_xls_2.pickle") or not os.path.exists("pickle\\contigs_bad_arachne_fasta_xls_2.pickle"):
        conts_xls = xlsManage.getArachneContigsOnScaffolds()                # int(cont_ID)  : (str(scaff_ID), int(start), int(end), int(cont_LEN))
        scaffs_fasta = fastaManage.loadScaffolds(fastaManage.ARACHNE)       # int(scaff_ID) : str(sekwencja)
        i = 0

        print "\nRozpoczynam wyszukiwanie contigow na scaffoldach wg danych z pliku new_data_arachne.xlsx"
        print "Contigow w XLSX do znalezienia w scaffoldach FASTA:", len(conts_xls)
        print "Scaffoldow w FASTA:", len(scaffs_fasta)
        for cont_xls_id, (scaff_xls_id, cont_xls_start, cont_xls_end, cont_xls_len) in conts_xls.iteritems():
            i += 1
            print "\rContig:", i,

            # ID contiga z pliku xlsx
            try:
                cont_id = int(cont_xls_id)
            except ValueError:
                print "Niepoprawny wpis w pliku new_data_arachne.xlsx. Bledna komorka - pusta!"
                continue

            # ID scaffoldu z pliku xlsx
            try:
                scaff_id = int(scaff_xls_id)
            except ValueError:
                scaff_id = int(str(scaff_xls_id)[:-1])

            # Czy scaffold o danym ID na ktorym lezy opisywany contig (wg xlsx) jest w pliku fasta i ma sekwencje?
            if not scaffs_fasta.has_key(scaff_id):
                print "Scaffold", scaff_id, "nie znaleziony w scaff_arachne.fasta!"
                continue
            else:
                scaff_seq = scaffs_fasta[scaff_id]
                
            # Sekwencja contiga
            try:
                cont_seq = scaff_seq[cont_xls_start:cont_xls_end]
                if len(cont_seq) != cont_xls_len:
                    print "Nie zgadza sie dlugosc contiga! (", len(cont_seq), "vs", cont_xls_len, ")", "; SCAFF_ID =", scaff_id, "--> (", cont_xls_start, ",", cont_xls_end, ")"
                    log = open(os.path.join('log', 'contig_arachne.log'), 'w')
                    log.write("SCAFFOLD ID:\t" + str(scaff_id) + "\n")
                    log.write("DLUGOSC SEKWENCJI SCAFFOLDU:\t" + str(len(scaff_seq)) + "\n")
                    log.write("CONTIG   ID:\t" + str(cont_id) + "\n")
                    log.write("\tPOSITION START:\t" + str(cont_xls_start) + "\n")
                    log.write("\tPOSITION STOP :\t" + str(cont_xls_end) + "\n")
                    log.write("\tDLUGOSC CONTIGA (XLSX):\t" + str(cont_xls_len) + "\n")
                    log.write("\n\n")
                    log.write("SEKWENCJA SCAFFOLDA:\n" + str(scaff_seq) + "\n")
                    log.write("\n\n")
                    log.write("SEKWENCJA CONTIGA:\n" + str(cont_seq) + "\n")
                    log.close()
                    sys.exit(1)
            except TypeError:
                print "\nCos nie tak ze START i STOP contiga o ID =", cont_id
                print "Start:", cont_xls_start
                print "Stop:", cont_xls_end
                sys.exit(1)
                
            # Mamy go na pewno - tworzymy tymczasowe obiekty
            contigs.append(ContigTemp(cont_id, scaff_id, cont_xls_start, cont_xls_end, cont_seq, cont_xls_len))

        print "\n\nZapisuje znalezione informacje w pliku pickle..."
        pickle.dump(contigs, open("pickle\\contigs_arachne_fasta_xls_2.pickle", "wb"))
        pickle.dump(bad_contigs, open("pickle\\contigs_bad_arachne_fasta_xls_2.pickle", "wb"))
    else:
        print "\nOdczytuje szczegolowe informacje o contigach z pliku pickle..."
        contigs = pickle.load(open("pickle\\contigs_arachne_fasta_xls.pickle", "rb"))
        bad_contigs = pickle.load(open("pickle\\contigs_bad_arachne_fasta_xls.pickle", "rb"))

    print "Tymczasowa baza zawiera", len(contigs), "contigow. Nie znaleziono:", len(bad_contigs), "contigow."

    return contigs, bad_contigs

##
#
#
def saveArachneContigsInDatabase(debug = False):
    contigs, _ = getContigsArachneFastaAndXLSX_2()
    
    contigs.sort(key = lambda x: x.start, reverse = False)
    contigs.sort(key = lambda x: x.cont_id, reverse = False)
    
    print "Ilosc contigow do zapisania:", len(contigs)
    
    for cont in contigs:
        ctg = Contig(id = cont.cont_id, 
                     scaff_id = cont.scaff_id,
                     order = -1,
                     start = cont.start,
                     end = cont.stop,
                     sequence = cont.sequence,
                     length_bp = cont.length)
        #ctg.save()


# =======================================================================================================================================
# ========================================== NIEZDEFINIOWANE CONTIGI CELERA =============================================================
# =======================================================================================================================================



# =======================================================================================================================================
# ========================================== NIEZDEFINIOWANE CONTIGI ARACHNE ============================================================
# =======================================================================================================================================




# =======================================================================================================================================
# ================================================== MARKERY CELERA =====================================================================
# =======================================================================================================================================

## Pomocnicza klasa Markera do przechowywania tymczasowych wartosci - bez zapisywania ich w bazie
class MarkerTemp:
    def __init__(self, name, chr_id, position, cont_id = -1, cont_start = -1, cont_stop = -1, scaff_id = -1, scaff_start = -1, scaff_stop = -1, sequence = "", length = -1, assemb = -1):
        self.name = name
        self.chr_id = chr_id
        self.position = position
        self.cont_id = cont_id
        self.cont_start = cont_start
        self.cont_stop = cont_stop
        self.scaff_id = scaff_id
        self.scaff_start = scaff_start
        self.scaff_stop = scaff_stop
        self.sequence = sequence
        self.length = length
        self.assemb = assemb

def getMarkersArachne():
    # Szukamy kazdego z markerow z xlsx na scaffoldach i contigach
    markers = []

    if not os.path.exists("pickle\\markers_arachne.pickle"):
        ### str(marker_name) : SLOWNIK(chr_id, position, contig_id, contig_start, contig_stop, scaff_id, scaff_start, scaff_stop)
        markers_xls = xlsManage.getArachneMarkers()
        ### str(marker_name) : str(sekwencja)
        markers_fasta = fastaManage.getArachneMarkers() 
        ### int(scaff_ID) : str(sekwencja)                    
        scaffs_fasta = fastaManage.loadScaffolds(fastaManage.ARACHNE)   
        ### int(cont_ID) : str(sekwencja)                    
        #conts_fasta = fastaManage.loadContigs(fastaManage.ARACHNE)    
        
        i = 0

        print "\nRozpoczynam wyszukiwanie markerow na scaffoldach i contigach wg danych z pliku new_data_arachne.xlsx"
        print "Markerow w XLSX do znalezienia:", len(markers_xls)
        print "Markerow w FASTA:", len(markers_fasta)
        print "Scaffoldow w FASTA:", len(scaffs_fasta)
        print "Contigow w FASTA:", 0

        for marker_name, marker_info in markers_xls.iteritems(): # (chr_id, position, contig_id, contig_start, contig_stop, scaff_id, scaff_start, scaff_stop)
            i += 1
            print "\rMarker:", i,
            
            # Sprawdzenie czy mamy sekwencje danego markera
            if not markers_fasta.has_key(marker_name):
                print "Brak sekwencji markera:", marker_name, "!"
                continue
            
            # Sprawdzenie czy istnieje scaffold o danym ID, na ktorym polozony jest dany marker
            if not scaffs_fasta.has_key(int(marker_info['scaff_id'])):
                print "Brak sekwencji scaffoldu, na ktorym podobno polozony jest marker:", marker_info['scaff_id']
                continue
            
            # Sprawdzenie czy dany marker rzeczywiscie znajduje sie na wlasciwym scaffoldzie
            if markers_fasta[marker_name] not in scaffs_fasta[int(marker_info['scaff_id'])]:
                print "Marker o nazwie", marker_name, "w rzeczywistosci nie znajduje sie na scaffoldzie o ID =", marker_info['scaff_id']
                continue

            # Sprawdzenie czy istnieje contig o danym ID, na ktorym polozony jest dany marker

            # Sprawdzenie czy dany marker rzeczywiscie znajduje sie na wlasciwym contigu



        #print "\n\nZapisuje znalezione informacje w pliku pickle..."
        #pickle.dump(markers, open("pickle\\markers_arachne.pickle", "wb"))
    else:
        print "\nOdczytuje szczegolowe informacje o markerach z pliku pickle..."
        markers = pickle.load(open("pickle\\markers_arachne.pickle", "rb"))

    print "\nTymczasowa baza zawiera", len(markers), "markerow."

    return markers

##
#
#
def saveArachneMarkersInDatabase(debug = False):
    markers = getMarkersArachne()
    
    markers.sort(key = lambda x: x.start, reverse = False)
    markers.sort(key = lambda x: x.cont_id, reverse = False)
    
    print "Ilosc markerow do zapisania:", len(markers)
    
    for mark in markers:
        # Zapisanie do bazy danych
        pass

# =======================================================================================================================================
# ================================================== MARKERY ARACHNE ====================================================================
# =======================================================================================================================================




# =======================================================================================================================================
# =======================================================================================================================================
# =======================================================================================================================================
# =======================================================================================================================================
# =======================================================================================================================================

def main(argv):
    try:
        opt, args = getopt.getopt(argv, "hcastud", [])
    except getopt.GetoptError, err:
        print str(err)
        usage()
        sys.exit(2)

    celera = False
    arachne = False
    scaffolds = False
    contigs = False
    undefined = False

    debug = False

    if len(opt) == 0:
        print "Brak parametrow wywolania!"
        usage()
        sys.exit(1)

    for o, a in opt:
        if o in ("-h", "--help"):
            usage()
            sys.exit()
        elif o == "-c":
            celera = True
        elif o == "-a":
            arachne = True
        elif o == "-s":
            scaffolds = True
        elif o == "-t":
            contigs = True
        elif o == "-u":
            undefined = True
        elif o == "-d":
            debug = True
        else:
            assert False, "Nieznana opcja!"

    if celera and scaffolds and undefined:
        print "Rozpoczynam zapis NIEZDEFINIOWANYCH scaffoldow CELERA w bazie danych..."
        if sure():
            saveCeleraUndefinedScaffoldsInDatabase("pickle\\scaffolds_celera.pickle", debug)
    elif arachne and scaffolds and undefined:
        pass
    elif celera and scaffolds:
        print "Rozpoczynam zapis ZDEFINIOWANYCH scaffoldow CELERA w bazie danych..."
        if sure():
            saveCeleraScaffoldsInDatabase("pickle\\scaffolds_celera.pickle", debug)
    elif arachne and scaffolds:
        print "Rozpoczynam zapis ZDEFINIOWANYCH scaffoldow ARACHNE w bazie danych..."
        if sure():
            saveArachneScaffoldsInDatabase("pickle\\scaffolds_arachne.pickle", debug)
    elif celera and contigs and undefined:
        pass
    elif arachne and contigs and undefined:
        pass
    elif celera and contigs:
        print "Rozpoczynam zapis ZDEFINIOWANYCH contigow CELERA w bazie danych..."
        if sure():
            saveCeleraContigsInDatabase(debug)
    elif arachne and contigs:
        print "Rozpoczynam zapis ZDEFINIOWANYCH contigow ARACHNE w bazie danych..."
        if sure():
            saveArachneContigsInDatabase(debug)
    else:
        print "Niepoprawne wywolanie programu!"
        usage()
        sys.exit(1)

def usage():
    usage = """
    Skrypt do wytworzenia bazy danych z plikow fasta oraz xlsx. Sposob uruchamiania:
 -h - pomoc                Wypisanie tekstu pomocy
 -c - celera
 -a - arachne
 -s - scaffolds
 -t - contigs
 -u - undefined
 -d - debug
    """
    print usage

def sure():
    while 1:
        if_sure = raw_input("\t\tCzy jestes pewny(-a)? (t / n) --> ")
        if if_sure == "t":
            return True
        elif if_sure == "n":
            return False
        else:
            continue

if __name__ == '__main__':
    main(sys.argv[1:])

    #saveCeleraScaffoldsInDatabase("pickle\\scaffolds_celera.pickle", True)
    #getContigsCeleraFasta()
    #checkCeleraContigs()
    #getContigsCeleraFastaAndXLSX()
    #printCeleraScaffoldsFromXLSX()
    #addLenAndSeqCeleraScaffolds()
    #checkScaffoldsSeqs()
    #getContigsArachneFastaAndXLSX_2()
    #getMarkersArachne()
    #saveArachneContigsInDatabase(True)