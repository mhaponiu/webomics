from openpyxl.reader.excel import load_workbook

# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# Funkcje bedace zaszloscia - prawdopodobnie beda niepotrzebne
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

def getCeleraScaffolds():
    xls_file = load_workbook("src\\laczenie060410.xlsx")
    print "Szukam..."
    for sheet in xls_file.worksheets:
        print "Arkusz: ", sheet.title
        if sheet.title != "finalne":
            continue
        # Arkusz finalne
        csv_file_name = "csv\\scaffolds_celera_id.csv"
        #csv_file_name = "csv\\scaffolds_arachne_id.csv"
        print "Tworze plik: %s, zawierajacy informacje o ID scaffoldu..." % csv_file_name
        csv_file = open(csv_file_name, "wt")
        # Ktore kolumny nas interesuja?
        first_row = sheet.rows[0]
        for i, cell in enumerate(first_row):
            #if "scf_celera" in cell.value:
            if "scf_ara_links_u" in cell.value:
                id_col = i
                continue
            #if "scf_cel_len" in cell.value:
            #    len_col = i
            if "chr_cel1" in cell.value:
                chr_col = i
        #print "Indeks ID: ", id_col, "; indeks dlugosci: ", len_col
        scaffolds_id = []
        for row in sheet.rows[1:]:
            scaffold = []
            # ID
            value = row[id_col].value
            if value in scaffolds_id:   # Mamy juz go!
                continue
            scaffolds_id.append(value)  # Odchaczamy, ze mamy
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)
            '''# DLUGOSC
            #value = row[len_col].value
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)'''
            # CHROMOSOM
            value = row[chr_col].value
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)
            # Zapisanie do CSV
            csv_file.write('\t'.join(scaffold))
            csv_file.write('\n')
        csv_file.close()

def getArachneScaffolds():
    xls_file = load_workbook("src\\Markery2.xlsx")
    print "Szukam..."
    for sheet in xls_file.worksheets:
        print "Arkusz: ", sheet.title
        if sheet.title != "Sheet1":
            continue
        csv_file_name = "csv\\scaffolds_arachne_id_from_markers.csv"
        print "Tworze plik: %s, zawierajacy informacje o ID scaffoldow arachne..." % csv_file_name
        csv_file = open(csv_file_name, "wt")
        # Ktore kolumny nas interesuja?
        first_row = sheet.rows[0]
        for i, cell in enumerate(first_row):
            #if "scf_celera" in cell.value:
            if "scaffold ara" in cell.value:
                id_col = i
                continue
            #if "scf_cel_len" in cell.value:
            #    len_col = i
            #if "chr_cel1" in cell.value:
            #    chr_col = i
        #print "Indeks ID: ", id_col, "; indeks dlugosci: ", len_col
        scaffolds_id = []
        for row in sheet.rows[1:]:
            scaffold = []
            # ID
            value = row[id_col].value
            if value in scaffolds_id:   # Mamy juz go!
                continue
            scaffolds_id.append(value)  # Odchaczamy, ze mamy
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)
            '''# DLUGOSC
            #value = row[len_col].value
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)'''
            '''# CHROMOSOM
            value = row[chr_col].value
            if value is None:
                value = ''
            if not isinstance(value, unicode):
                value = unicode(value)
            value = value.encode('utf8')
            scaffold.append(value)'''
            # Zapisanie do CSV
            csv_file.write('\t'.join(scaffold))
            csv_file.write('\n')
        csv_file.close()

#getCeleraScaffolds()
#getArachneScaffolds()


