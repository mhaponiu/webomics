import sys
import os
import pickle

from openpyxl.reader.excel import load_workbook

this_module_path = os.path.abspath (os.path.dirname(sys.modules[__name__].__file__))
xlsx_src_path = os.path.join(this_module_path, "src")
pickle_path = os.path.join(this_module_path, "pickle")

laczenie_file_path = os.path.join(xlsx_src_path, "laczenie.xlsx")
markers_file_path = os.path.join(xlsx_src_path, "Markery2.xlsx")

ARACHNE = 0
CELERA = 1

SCAFFOLD = 0
CONTIG = 1

# =======================================================================================================================================
# ======================================================== SCAFFOLDY ====================================================================
# =======================================================================================================================================

## Funkcja tworzaca scaffoldy celera na podstawie pliku laczenie i arkuszy chr1_a...chr1_n.
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Zwraca umieszczenie scaffoldow celera na chromosomach (slownik) wraz z pozycjami kolejnych scaffoldow
def getCeleraScaffolds(debug = False):
    chromosomes = {}
    
    xls_file = load_workbook(laczenie_file_path)
    
    for sheet in xls_file.worksheets:
        if sheet.title in ['chr1_a', 'chr2_a', 'chr3_a', 'chr4_a', 'chr5_a', 'chr6_a', 'chr7_a']:
            if debug == True:
                print "Arkusz: ", sheet.title
            scaffolds = {}
            # Odczytanie pozycji kolumn
            first_row = sheet.rows[0]
            scaff_id_col = -1
            scaff_pos_col = -1
            for i, cell in enumerate(first_row):
                # Odczytanie ID scaffoldu
                if "scf_celera" in cell.value:
                    scaff_id_col = i
                    continue
                if "cM" in cell.value:
                    scaff_pos_col = i
                    continue
            if scaff_id_col == -1:
                raise "Nie znaleziono kolumny 'scf_celera'!"
                sys.exit(1)
            if scaff_pos_col == -1:
                raise "Nie znaleziono kolumny 'cM'!"
                sys.exit(1)

            last_id = -1
            order = 0
            # Kolejne scaffoldy
            for row in sheet.rows[1:]:
                try:
                    # ID
                    value_id = int(row[scaff_id_col].value)
                    if not isinstance(value_id, unicode):
                        value_id = unicode(value_id)
                    value_id = value_id.encode('utf8')
                    if not scaffolds.has_key(value_id):     # Nie mamy jeszcze tego ID
                        scaffolds[value_id] = []            # Inicjalizujemy pusta tablice
                except ValueError:
                    if debug == True:
                        print "Scaffold dzielony:", row[scaff_id_col].value
                    value_id = row[scaff_id_col].value
                    if not isinstance(value_id, unicode):
                        value_id = unicode(value_id)
                    value_id = value_id.encode('utf8')
                    if not scaffolds.has_key(value_id):     # Nie mamy jeszcze tego ID
                        scaffolds[value_id] = []            # Inicjalizujemy pusta tablice

                # POZYCJA
                value_pos = float(row[scaff_pos_col].value)
                if not isinstance(value_pos, unicode):
                    value_pos = unicode(value_pos)
                value_pos = value_pos.encode('utf8')

                if value_id == last_id:
                    scaffolds[value_id][-1][1].append(value_pos)   # Dodajemy do istniejacego fragmentu scaffolda - ostatnio dodanego
                else:
                    scaffolds[value_id].append((order, [value_pos]))                    # Inicjujemy nowy fragment scaffolda i dodajemy jego pierwszy element
                    order += 1

                last_id = value_id

            chromosomes[int(str(sheet.title)[3])] = scaffolds
    
    return chromosomes

## Funkcja tworzaca scaffoldy arachne na podstawie pliku Markery.
#  @param debug: True oznacza wypisywanie dodatkowych informacji.
#  @return: Zwraca umieszczenie scaffoldow arachne na chromosomach (slownik) wraz z pozycjami kolejnych scaffoldow
def getArachneScaffolds(debug = False):
    chromosomes = {}
    
    xls_file = load_workbook(markers_file_path)
    
    for sheet in xls_file.worksheets:
        if sheet.title == 'Sheet1':
            if debug == True:
                print "Arkusz: ", sheet.title, "..."
            # Odczytanie pozycji kolumn
            first_row = sheet.rows[0]
            chr_id_col = -1
            scaff_id_col = -1
            scaff_pos_col = -1
            for i, cell in enumerate(first_row):
                # Odczytanie ID chromosomu
                if "chromosom" in cell.value:
                    chr_id_col = i
                    continue
                # Odczytanie ID scaffoldu
                if "scaffold ara" in cell.value:
                    scaff_id_col = i
                    continue
                # Odczytanie pozycji scaffoldu
                if "pozycja cM" in cell.value:
                    scaff_pos_col = i
                    continue
            if chr_id_col == -1:
                raise "Nie znaleziono kolumny 'chromosom'!"
                sys.exit(1)    
            if scaff_id_col == -1:
                raise "Nie znaleziono kolumny 'scf_celera'!"
                sys.exit(1)
            if scaff_pos_col == -1:
                raise "Nie znaleziono kolumny 'cM'!"
                sys.exit(1)

            last_id = -1
            order = 0
            # Kolejne wiersze
            for row in sheet.rows[1:]:
                chr_id = int(str((row[chr_id_col].value))[3])
                # Nie mamy jeszcze tego chromosoma
                if not chromosomes.has_key(chr_id):
                    chromosomes[chr_id] = {}    # Inicjalizujemy
                
                try:
                    # ID scaffolda
                    value_id = int(row[scaff_id_col].value)
                    if not chromosomes[chr_id].has_key(value_id):     # Nie mamy jeszcze tego ID
                        chromosomes[chr_id][value_id] = []            # Inicjalizujemy pusta tablice
                except ValueError:
                    print "[X] ID nie jest liczba calkowita!"
                    sys.exit(1)

                # POZYCJA
                value_pos = float(row[scaff_pos_col].value)

                if value_id == last_id:
                    chromosomes[chr_id][value_id][-1][1].append(value_pos)      # Dodajemy do istniejacego fragmentu scaffolda - ostatnio dodanego
                else:
                    chromosomes[chr_id][value_id].append((order, [value_pos]))  # Inicjujemy nowy fragment scaffolda i dodajemy jego pierwszy element
                    order += 1

                last_id = value_id
    
    return chromosomes

## Funkcja zwraca dlugosci scaffoldow (slownik) odczytanych z pliku laczenie
def getScaffoldsLenFromXLS(assem, debug = False):
    xls_file = load_workbook(laczenie_file_path)
    scaffolds = {}
    for sheet in xls_file.worksheets:
        if sheet.title in ['finalne']:
            # Odczytanie pozycji kolumn
            first_row = sheet.rows[0]
            scaff_id_col = -1
            scaff_len_col = -1
            # !!!
            # CELERA
            # !!!
            if assem == CELERA:
                for i, cell in enumerate(first_row):
                    # Odczytanie ID scaffoldu
                    if "scf_celera" in cell.value:
                        scaff_id_col = i
                    # Odczytanie dlugosci scaffoldu
                    if "scf_cel_len" in cell.value:
                        scaff_len_col = i
                        continue
                if scaff_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_celera'!"
                    sys.exit(1)
                if scaff_len_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_cel_len'!"
                    sys.exit(1)

                # Kolejne scaffoldy
                for row in sheet.rows[1:]:
                    try:
                        value_id = str(row[scaff_id_col].value)
                        if not isinstance(value_id, unicode):
                            value_id = unicode(value_id)
                        value_id = value_id.encode('utf8')
                        if not scaffolds.has_key(value_id):     # Nie mamy jeszcze tego ID
                            scaffolds[value_id] = str(row[scaff_len_col].value)
                        else:
                            continue
                    except ValueError:
                        if debug == True:
                            print "[X] Scaffold dzielony:", row[scaff_id_col].value
                        value_id = row[scaff_id_col].value
                        if not isinstance(value_id, unicode):
                            value_id = unicode(value_id)
                        value_id = value_id.encode('utf8')
                        if not scaffolds.has_key(value_id):     # Nie mamy jeszcze tego ID
                            scaffolds[value_id] = str(row[scaff_len_col].value)
                        else:
                            continue
            # !!!
            # ARACHNE
            # !!!
            elif assem == ARACHNE:
                for i, cell in enumerate(first_row):
                    # Odczytanie ID scaffoldu
                    if "scf_ara_links" in cell.value:
                        scaff_id_col = i
                    # Odczytanie dlugosci scaffoldu
                    if "scf_ara_dlugosc" in cell.value:
                        scaff_len_col = i
                        continue
                if scaff_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_ara_links'!"
                    sys.exit(1)
                if scaff_len_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_ara_dlugosc'!"
                    sys.exit(1)

                # Kolejne scaffoldy
                for row in sheet.rows[1:]:
                    try:
                        value_id = str(row[scaff_id_col].value)
                        if not isinstance(value_id, unicode):
                            value_id = unicode(value_id)
                        value_id = value_id.encode('utf8')
                        if not scaffolds.has_key(value_id):     # Nie mamy jeszcze tego ID
                            scaffolds[value_id] = str(row[scaff_len_col].value)
                        else:
                            continue
                    except ValueError:
                        if debug == True:
                            print "[X] Blad w wierszu z ID:", row[scaff_id_col].value, "\twartosc:", row[scaff_len_col].value
                        continue
            else:
                print "[X] Niepoprawny typ asemblacji!"
                sys.exit(2)

    return scaffolds

## Funkcja zwraca dlugosci scaffoldow pobranych z pliku laczenie - jezeli istnieje plik pickle - nie analizuje arkusza xlsx - odczytuje plik pickle i jego zwraca
#  ID - klucz zwracanego slownika - typ string, wartosc klucza - typ string
def getScaffoldsLen(assem, debug = False):
    scaffolds = {}

    # !!!
    # CELERA
    # !!!
    if assem == CELERA:
        if not os.path.exists(os.path.join(pickle_path, "scaffolds_len_celera.pickle")):
            print "\n\n[X] Pobieram dlugosci scaffoldow CELERA z pliku XLSX (laczenie)..."
            scaffolds = getScaffoldsLenFromXLS(CELERA, debug)
            print "\n[X] Zapisuje znalezione scaffoldy w pliku pickle\\scaffolds_len_celera.pickle..."
            pickle.dump(scaffolds, open(os.path.join(pickle_path, "scaffolds_len_celera.pickle"), "wb"))
        else:
            print "\n[X] Odczytuje dlugosci scaffoldow CELERA z pliku pickle..."
            scaffolds = pickle.load(open(os.path.join(pickle_path, "scaffolds_len_celera.pickle"), "rb"))
    # !!!
    # ARACHNE
    # !!!
    elif assem == ARACHNE:
        if not os.path.exists(os.path.join(pickle_path, "scaffolds_len_arachne.pickle")):
            print "\n\n[X] Pobieram dlugosci scaffoldow ARACHNE z pliku XLSX (laczenie)..."
            scaffolds = getScaffoldsLenFromXLS(ARACHNE, debug)
            print "\n[X] Zapisuje znalezione scaffoldy w pliku pickle\\scaffolds_len_arachne.pickle..."
            pickle.dump(scaffolds, open(os.path.join(pickle_path, "scaffolds_len_arachne.pickle"), "wb"))
        else:
            print "\n[X] Odczytuje dlugosci scaffoldow ARACHNE z pliku pickle..."
            scaffolds = pickle.load(open(os.path.join(pickle_path, "scaffolds_len_arachne.pickle"), "rb"))
    else:
        print "[X] Niepoprawny typ asemblacji!"
        sys.exit(2)

    print "[X] Liczba odczytanych scaffoldow:", len(scaffolds)

    return scaffolds

## Funkcja zwraca dlugosc scaffolda o zadanym ID
#  @param assemb: Typ asemblacji
#  @param scaff_id: ID scaffoldu, ktorego dlugosci poszukujemy.
#  @return: Dlugosc scaffoldu o zadanym ID
def getScaffLen(assemb, scaff_id):
    scaffs = getScaffoldsLen(assemb, False)
    return scaffs[str(scaff_id)]

## Funkcja pobiera z pliku laczenie informacje na temat przynaleznosci contigow celery do scaffoldow celery
#  @return: Zwraca slownik: contigs[ID_contiga] = tupla(ID_scaffoldu, dlugosc_contiga)
def getCeleraContigsOnScaffolds():
    contigs = {}
    if os.path.exists(os.path.join(pickle_path, "contigs_scaffolds_celera.pickle")):
        print "\n[X] Odczytuje przyporzadkowanie contigow CELERA do scaffoldow z pliku pickle..."
        contigs = pickle.load(open(os.path.join(pickle_path, "contigs_scaffolds_celera.pickle"), "rb"))
    else:
        print "\n\n[X] Pobieram przyporzadkowanie contigow CELERA do scaffoldow z pliku XLSX (laczenie)..."
        xls_file = load_workbook(os.path.join(xlsx_src_path, "laczenie.xlsx"))
        for sheet in xls_file.worksheets:
            if sheet.title in ['finalne']:
                # Odczytanie pozycji kolumn
                first_row = sheet.rows[0]
                cont_id_col = -1
                scaff_id_col = -1
                cont_len_col = -1
                for i, cell in enumerate(first_row):
                    # Odczytanie ID contigu
                    if "celera" == cell.value:
                        cont_id_col = i
                        continue
                    # Odczytanie dlugosci contigu
                    if "cel_ctg_len" == cell.value:
                        cont_len_col = i
                        continue
                    # Odczytanie ID scaffoldu
                    if "scf_celera" in cell.value:
                        scaff_id_col = i
                        continue
                if cont_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'celera'!"
                    sys.exit(1)
                if scaff_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_celera'!"
                    sys.exit(1)

                # Kolejne contigi
                for row in sheet.rows[1:]:
                    cont_id = int(row[cont_id_col].value)
                    if not isinstance(cont_id, unicode):
                        cont_id = unicode(cont_id)
                    cont_id = cont_id.encode('utf8')
                    if not contigs.has_key(cont_id):     # Nie mamy jeszcze tego ID
                        contigs[cont_id] = (str(row[scaff_id_col].value), str(row[cont_len_col].value))
                    else:
                        continue

        print "\n[X] Zapisuje przyporzadkowanie contigow CELERA do scaffoldow w pliku pickle\\contigs_scaffolds_celera.pickle..."
        pickle.dump(contigs, open(os.path.join(pickle_path, "contigs_scaffolds_celera.pickle"), "wb"))

    print "\n[X] Odczytano", len(contigs), "contigow i ich przyporzadkowan do scaffoldow."
    return contigs

## Funkcja pobiera z pliku Markery informacje na temat przynaleznosci contigow arachne do scaffoldow arachne
#  @return: Zwraca slownik: contigs[ID_contiga] = tupla(ID_scaffoldu, dlugosc_contiga)
def getArachneContigsOnScaffolds():
    contigs = {}
    if os.path.exists(os.path.join(pickle_path, "contigs_scaffolds_arachne.pickle")):
        print "\n[X] Odczytuje przyporzadkowanie contigow ARACHNE do scaffoldow z pliku pickle..."
        contigs = pickle.load(open(os.path.join(pickle_path, "contigs_scaffolds_arachne.pickle"), "rb"))
    else:
        print "\n\n[X] Pobieram przyporzadkowanie contigow ARACHNE do scaffoldow z pliku XLSX (new_data_arachne)..."
        xls_file = load_workbook(os.path.join(xlsx_src_path, "new_data_arachne.xlsx"))
        for sheet in xls_file.worksheets:
            if sheet.title in ['Scf+ctg']:
                # Odczytanie pozycji kolumn
                first_row = sheet.rows[0]
                cont_id_col = -1
                scaff_id_col = -1
                cont_len_col = -1
                cont_start_col = -1
                cont_end_col = -1
                for i, cell in enumerate(first_row):
                    # Odczytanie ID contigu
                    if "ctg_id" == cell.value:
                        cont_id_col = i
                        continue
                    # Odczytanie dlugosci contigu
                    if "length_of_ctg" == cell.value:
                        cont_len_col = i
                        continue
                    # Odczytanie ID scaffoldu
                    if "scf_id" in cell.value:
                        scaff_id_col = i
                        continue
                    # Odczytanie pozycji START contigu
                    if "start_contig" == cell.value:
                        cont_start_col = i
                        continue
                    # Odczytanie END contigu
                    if "stop_contig" in cell.value:
                        cont_end_col = i
                        continue
                if cont_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'ctg_id'!"
                    sys.exit(1)
                if scaff_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scf_id'!"
                    sys.exit(1)
                if cont_len_col == -1:
                    raise "[X] Nie znaleziono kolumny 'length_of_ctg'!"
                    sys.exit(1)
                if cont_start_col == -1:
                    raise "[X] Nie znaleziono kolumny 'start_contig'!"
                    sys.exit(1)
                if cont_end_col == -1:
                    raise "[X] Nie znaleziono kolumny 'stop_contig'!"
                    sys.exit(1)    

                # Kolejne contigi
                for row in sheet.rows[1:]:
                    # CONTIG ID
                    cont_id = int(row[cont_id_col].value)
                    
                    # SCAFFOLD ID
                    scaff_id = str(row[scaff_id_col].value)
                    
                    # CONTIG LENGTH
                    cont_len = int(row[cont_len_col].value)
                    
                    # CONTIG START
                    cont_start = int(row[cont_start_col].value)
                    
                    # CONTIG END
                    cont_end = int(row[cont_end_col].value)
                    
                    if not contigs.has_key(cont_id):     # Nie mamy jeszcze tego ID
                        contigs[cont_id] = (scaff_id, cont_start, cont_end, cont_len)
                    else:
                        continue

        print "\n[X] Zapisuje przyporzadkowanie contigow ARACHNE do scaffoldow w pliku pickle\\contigs_scaffolds_arachne.pickle..."
        pickle.dump(contigs, open(os.path.join(pickle_path, "contigs_scaffolds_arachne.pickle"), "wb"))

    print "\n[X] Odczytano", len(contigs), "contigow i ich przyporzadkowan do scaffoldow."
    return contigs

### MARKERY

def getArachneMarkers():
    markers = {}
    if os.path.exists(os.path.join(pickle_path, "markers_arachne.pickle")):
        print "\n[X] Odczytuje przyporzadkowanie markerow ARACHNE do scaffoldow i contigow z pliku pickle..."
        markers = pickle.load(open(os.path.join(pickle_path, "markers_arachne.pickle"), "rb"))
    else:
        print "\n\n[X] Pobieram przyporzadkowanie markerow ARACHNE do scaffoldow i contigow z pliku XLSX (new_data_arachne)..."
        xls_file = load_workbook(os.path.join(xlsx_src_path, "new_data_arachne.xlsx"))
        for sheet in xls_file.worksheets:
            if sheet.title in ['Markery']:
                # Odczytanie pozycji kolumn
                first_row = sheet.rows[0]
                chr_id_col = -1
                marker_name_col = -1
                position_cm_col = -1
                contig_id_col = -1
                contig_start_col = -1
                contig_stop_col = -1
                scaff_id_col = -1
                scaff_start_col = -1
                scaff_stop_col = -1
            
                for i, cell in enumerate(first_row):
                    # Odczytanie ID chromosomu
                    if "chr_id" == cell.value:
                        chr_id_col = i
                        continue
                    # Odczytanie nazwy markera
                    if "marker" == cell.value:
                        marker_name_col = i
                        continue
                    # Odczytanie pozycji markera w cM
                    if "pozycja cM" == cell.value:
                        position_cm_col = i
                        continue
                    # Odczytanie ID contiga
                    if "kontig nr" == cell.value:
                        contig_id_col = i
                        continue
                    # Odczytanie pozycji start na contigu
                    if "kontig start" == cell.value:
                        contig_start_col = i
                        continue
                    # Odczytanie pozycji stop na contigu
                    if "kontig stop" == cell.value:
                        contig_stop_col = i
                        continue
                    # Odczytanie ID scaffolda
                    if "scaffold nr" == cell.value:
                        scaff_id_col = i
                        continue
                    # Odczytanie pozycji start na scaffoldzie
                    if "scaffold start" == cell.value:
                        scaff_start_col = i
                        continue
                    # Odczytanie pozycji stop na scaffoldzie
                    if "scaffold stop" == cell.value:
                        scaff_stop_col = i
                        continue
                    
                if chr_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'chr_id'!"
                    sys.exit(1)
                if marker_name_col == -1:
                    raise "[X] Nie znaleziono kolumny 'marker'!"
                    sys.exit(1)
                if position_cm_col == -1:
                    raise "[X] Nie znaleziono kolumny 'pozycja cM'!"
                    sys.exit(1)   
                if contig_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'kontig nr'!"
                    sys.exit(1)  
                if contig_start_col == -1:
                    raise "[X] Nie znaleziono kolumny 'kontig start'!"
                    sys.exit(1)     
                if contig_stop_col == -1:
                    raise "[X] Nie znaleziono kolumny 'kontig stop'!"
                    sys.exit(1)    
                if scaff_id_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scaffold nr'!"
                    sys.exit(1)      
                if scaff_start_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scaffold start'!"
                    sys.exit(1)     
                if scaff_stop_col == -1:
                    raise "[X] Nie znaleziono kolumny 'scaffold stop'!"
                    sys.exit(1)      
                          
                # Kolejne markery
                for row in sheet.rows[1:]:
                    # CHROMOSOME ID
                    chr_id = int(row[chr_id_col].value)
                    # MARKER NAME
                    marker_name = str(row[marker_name_col].value)
                    # POSITION_CM
                    pocition_cm = float(row[position_cm_col].value)
                    # CONTIG ID
                    contig_id = str(row[contig_id_col].value)
                    # CONTIG START
                    contig_start = str(row[contig_start_col].value)
                    # CONTIG STOP
                    contig_stop = str(row[contig_stop_col].value)
                    # SCAFFOLD ID
                    scaffold_id = str(row[scaff_id_col].value)
                    # SCAFFOLD START
                    scaffold_start = str(row[scaff_start_col].value)
                    # SCAFFOLD STOP
                    scaffold_stop = str(row[scaff_stop_col].value)
                    
                    if not markers.has_key(marker_name):     # Nie mamy jeszcze tego markera
                        markers[marker_name] = {"chr_id" : chr_id, 
                                                "position" : pocition_cm, 
                                                "contig_id" : contig_id, 
                                                "contig_start" : contig_start, 
                                                "contig_stop" : contig_stop, 
                                                "scaff_id" : scaffold_id, 
                                                "scaff_start" : scaffold_start, 
                                                "scaff_stop" : scaffold_stop}
                    else:
                        print "Mamy juz markera:", marker_name
                        continue

        print "\n[X] Zapisuje przyporzadkowanie contigow ARACHNE do scaffoldow w pliku pickle\\markers_arachne.pickle..."
        pickle.dump(markers, open(os.path.join(pickle_path, "markers_arachne.pickle"), "wb"))

    print "\n[X] Odczytano", len(markers), "markerow i ich przyporzadkowan do scaffoldow oraz contigow."
    return markers


def main():
    #print getScaffLen(CELERA, 360114)

    #chroms = getArachneScaffolds(True)
    #chroms = getCeleraScaffolds(True)
    
    #for k, v in chroms[1].iteritems():
    #    print k, v
    
    conts = getArachneContigsOnScaffolds()  # int(cont_ID)  : (str(scaff_ID), int(start), int(end), int(cont_LEN))
    
    #for cont_id, (scaff_id, cont_start, cont_end, cont_len) in conts.iteritems():
    #    print "ID =", cont_id, "-->", scaff_id, "-->", cont_start, ";", cont_end, "-->", cont_len
    
if __name__ == "__main__":
    main()
